from __future__ import annotations

import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from wastewater_snd.schema import (
    AERATION_COL,
    A_LIVE_COL,
    A_MAX_COL,
    COD_IN_COL,
    DATE_COL,
    H_LIVE_COL,
    H_MAX_COL,
    MODEL_DRAFT_COLUMNS,
    N_LIVE_COL,
    N_MAX_COL,
    OPTIONAL_MODEL_COLUMNS,
    REMOVAL_COL,
    REQUIRED_MODEL_COLUMNS,
    SND_COL,
    TEMP_COL,
    TN_IN_COL,
    TN_OUT_COL,
    QualityIssue,
)


MODEL_NUMERIC_COLUMNS = [
    column for column in REQUIRED_MODEL_COLUMNS + OPTIONAL_MODEL_COLUMNS if column != DATE_COL
]


def _clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _number(value: object) -> float:
    result = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return float(result) if pd.notna(result) else math.nan


def _month_day_number(
    value: float,
    year: int,
    candidates: set[str] | None = None,
) -> date:
    month = int(math.floor(value))
    if candidates:
        matches: list[date] = []
        for candidate in candidates:
            parsed = date.fromisoformat(candidate)
            collapsed = float(f"{parsed.month}.{parsed.day}")
            if parsed.year == year and math.isclose(collapsed, value, abs_tol=1e-9):
                matches.append(parsed)
        if len(matches) == 1:
            return matches[0]
    fraction = value - month
    # Excel collapses both M.D and M.DD to a float. Prefer two decimal places
    # (5.2 -> May 20 in these sheets); if that is not a calendar day, fall back
    # to one decimal place (5.6 -> May 6, because day 60 is impossible).
    day = int(round(fraction * 100))
    try:
        return date(year, month, day)
    except ValueError:
        return date(year, month, int(round(fraction * 10)))


def canonical_date(
    value: object,
    year: int = 2026,
    candidates: set[str] | None = None,
) -> str | None:
    """Convert Excel dates and the project's M.DD notation to ISO dates.

    The workbooks store values such as 5.20 as the number 5.2. Multiplying the
    fractional part by 100 intentionally restores day 20 rather than day 2.
    """
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    if isinstance(value, (int, float, np.integer, np.floating)):
        numeric = float(value)
        if numeric >= 10_000:
            parsed = pd.Timestamp("1899-12-30") + pd.to_timedelta(numeric, unit="D")
            return parsed.strftime("%Y-%m-%d")
        return _month_day_number(numeric, year, candidates=candidates).isoformat()

    text = _clean_text(value)
    if not text:
        return None
    try:
        numeric = float(text)
    except ValueError:
        numeric = math.nan
    if not math.isnan(numeric):
        return canonical_date(numeric, year=year, candidates=candidates)

    normalized = text.replace("/", ".").replace("-", ".")
    parts = [part for part in normalized.split(".") if part]
    if len(parts) == 3 and len(parts[0]) == 4:
        return date(int(parts[0]), int(parts[1]), int(parts[2])).isoformat()
    if len(parts) == 2:
        return date(year, int(parts[0]), int(parts[1])).isoformat()
    parsed = pd.to_datetime(text, errors="coerce")
    return None if pd.isna(parsed) else parsed.strftime("%Y-%m-%d")


def _parse_sample_label(
    value: object,
    year: int = 2026,
    candidates: set[str] | None = None,
) -> tuple[str, str] | None:
    text = _clean_text(value).upper()
    match = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)([JC])", text)
    if not match:
        return None
    parsed_date = canonical_date(
        float(match.group(1)), year=year, candidates=candidates
    )
    if parsed_date is None:
        return None
    return parsed_date, match.group(2)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _workbook_metadata(path: Path) -> dict[str, object]:
    book = pd.ExcelFile(path)
    sheets: dict[str, object] = {}
    for sheet_name in book.sheet_names:
        frame = pd.read_excel(path, sheet_name=sheet_name, header=None)
        sheets[sheet_name] = {
            "rows": int(frame.shape[0]),
            "columns": int(frame.shape[1]),
            "nonempty_rows": int(frame.notna().any(axis=1).sum()),
        }
    return {
        "path": str(path.resolve()),
        "size_bytes": path.stat().st_size,
        "sha256": _sha256(path),
        "sheets": sheets,
    }


def _formula_issues(path: Path) -> list[QualityIssue]:
    issues: list[QualityIssue] = []
    workbook = load_workbook(path, read_only=True, data_only=False)
    markers = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(marker in value for marker in markers):
                    issues.append(
                        QualityIssue(
                            "error",
                            "BROKEN_EXCEL_FORMULA",
                            path.name,
                            sheet.title,
                            cell.row,
                            f"{cell.coordinate} 包含失效公式：{value}",
                        )
                    )
    return issues


def _parse_in_out_sheet(
    path: Path,
    sheet: str,
    value_col: int,
    label_col: int,
    removal_col: int | None,
    date_candidates: set[str] | None = None,
) -> tuple[pd.DataFrame, list[QualityIssue]]:
    frame = pd.read_excel(path, sheet_name=sheet, header=None)
    issues: list[QualityIssue] = []
    records: list[dict[str, object]] = []
    current_in: dict[str, float] = {}
    current_in_row: dict[str, int] = {}
    out_counts: Counter[str] = Counter()
    inflow_counts: Counter[str] = Counter()
    last_in_date: str | None = None

    for index in range(1, len(frame)):
        parsed = _parse_sample_label(
            frame.iat[index, label_col], candidates=date_candidates
        )
        if parsed is None:
            continue
        sample_date, role = parsed
        value = _number(frame.iat[index, value_col])
        removal = (
            _number(frame.iat[index, removal_col]) if removal_col is not None else math.nan
        )

        if role == "J" and not math.isnan(removal):
            issues.append(
                QualityIssue(
                    "warning",
                    "ROLE_INFERRED_FROM_REMOVAL",
                    path.name,
                    sheet,
                    index + 1,
                    "标签为 J，但去除率非空；按出水 C 解析，原标签保留在问题清单。",
                )
            )
            role = "C"

        if role == "J":
            inflow_counts[sample_date] += 1
            if inflow_counts[sample_date] > 1:
                issues.append(
                    QualityIssue(
                        "warning",
                        "DUPLICATE_INFLOW_DATE",
                        path.name,
                        sheet,
                        index + 1,
                        f"{sample_date} 出现第 {inflow_counts[sample_date]} 条进水记录。",
                    )
                )
            current_in[sample_date] = value
            current_in_row[sample_date] = index + 1
            last_in_date = sample_date
            continue

        if last_in_date is not None and sample_date != last_in_date:
            issues.append(
                QualityIssue(
                    "warning",
                    "OUTFLOW_DATE_DIFFERS_FROM_PRECEDING_INFLOW",
                    path.name,
                    sheet,
                    index + 1,
                    f"出水标签 {sample_date} 与最近进水标签 {last_in_date} 不一致；未自动改日期。",
                )
            )
        out_counts[sample_date] += 1
        records.append(
            {
                "date": sample_date,
                "source_sample_index": int(out_counts[sample_date]),
                "in_value": current_in.get(sample_date, math.nan),
                "out_value": value,
                "removal_rate": removal,
                "source_row": index + 1,
                "inflow_source_row": current_in_row.get(sample_date),
            }
        )

    result = pd.DataFrame.from_records(records)
    return result, issues


def _parse_snd_sheet(
    path: Path,
    sheet: str,
    date_candidates: set[str] | None = None,
) -> tuple[pd.DataFrame, list[QualityIssue]]:
    frame = pd.read_excel(path, sheet_name=sheet, header=None)
    issues: list[QualityIssue] = []
    records: list[dict[str, object]] = []
    counts: Counter[str] = Counter()
    names = [
        "no3_in_mg_l",
        "no3_out_mg_l",
        "no2_in_mg_l",
        "no2_out_mg_l",
        "nh4_in_mg_l",
        "nh4_out_mg_l",
        "snd_rate",
    ]
    for index in range(1, len(frame)):
        sample_date = canonical_date(
            frame.iat[index, 0], candidates=date_candidates
        )
        if sample_date is None:
            continue
        counts[sample_date] += 1
        values = [_number(frame.iat[index, column]) for column in range(1, 8)]
        record: dict[str, object] = {
            "date": sample_date,
            "source_sample_index": int(counts[sample_date]),
            "source_row": index + 1,
        }
        record.update(dict(zip(names, values, strict=True)))
        records.append(record)
        snd_value = values[-1]
        if math.isnan(snd_value) or not 0 <= snd_value <= 1:
            issues.append(
                QualityIssue(
                    "error",
                    "INVALID_SND_RATE",
                    path.name,
                    sheet,
                    index + 1,
                    f"SND 率应位于 0–1，当前值为 {snd_value!r}。",
                )
            )
    return pd.DataFrame.from_records(records), issues


def _scan_negative_analytical_results(path: Path) -> list[QualityIssue]:
    issues: list[QualityIssue] = []
    book = pd.ExcelFile(path)
    specifications = {
        "硝氮": (6, "NEGATIVE_NO3_RESULT"),
        "亚硝氮": (3, "NEGATIVE_NO2_RESULT"),
    }
    for sheet, (column, code) in specifications.items():
        if sheet not in book.sheet_names:
            continue
        frame = pd.read_excel(path, sheet_name=sheet, header=None)
        for index in range(1, len(frame)):
            value = _number(frame.iat[index, column])
            if not math.isnan(value) and value < 0:
                issues.append(
                    QualityIssue(
                        "warning",
                        code,
                        path.name,
                        sheet,
                        index + 1,
                        f"检测结果为 {value:.6g} mg/L；应按方法检出限标记，不应直接当普通负值或零值建模。",
                    )
                )
    return issues


def parse_water_quality(
    path: Path,
    *,
    legacy: bool,
) -> tuple[pd.DataFrame, list[QualityIssue], dict[str, int]]:
    issues = _formula_issues(path) + _scan_negative_analytical_results(path)
    if legacy:
        tn_sheet, cod_sheet, nh4_sheet, snd_sheet = "TN", "COD", "氨氮", "SND"
        tn_spec = (0, 1, 2)
        cod_spec = (0, 3, 5)
        nh4_spec = (0, 3, 4)
    else:
        tn_sheet, cod_sheet, nh4_sheet, snd_sheet = "总氮", "COD", "氨氮", "SND率"
        tn_spec = (0, 1, 2)
        cod_spec = (0, 3, 5)
        nh4_spec = (0, 1, 2)

    nitrate = pd.read_excel(path, sheet_name="硝氮", header=None)
    date_candidates = {
        parsed
        for value in nitrate.iloc[1:, 7]
        if (parsed := canonical_date(value)) is not None
    }
    tn, tn_issues = _parse_in_out_sheet(
        path, tn_sheet, *tn_spec, date_candidates=date_candidates
    )
    cod, cod_issues = _parse_in_out_sheet(
        path, cod_sheet, *cod_spec, date_candidates=date_candidates
    )
    nh4, nh4_issues = _parse_in_out_sheet(
        path, nh4_sheet, *nh4_spec, date_candidates=date_candidates
    )
    snd, snd_issues = _parse_snd_sheet(
        path, snd_sheet, date_candidates=date_candidates
    )
    issues.extend(tn_issues + cod_issues + nh4_issues + snd_issues)

    key = ["date", "source_sample_index"]
    water = snd.copy()
    water = water.merge(
        tn[key + ["in_value", "out_value", "removal_rate"]].rename(
            columns={
                "in_value": "tn_in_mg_l",
                "out_value": "tn_out_mg_l",
                "removal_rate": "tn_removal_rate",
            }
        ),
        how="left",
        on=key,
    )
    water = water.merge(
        cod[key + ["in_value", "out_value", "removal_rate"]].rename(
            columns={
                "in_value": "cod_in_mg_l",
                "out_value": "cod_out_mg_l",
                "removal_rate": "cod_removal_rate",
            }
        ),
        how="left",
        on=key,
    )
    water = water.merge(
        nh4[key + ["in_value", "out_value", "removal_rate"]].rename(
            columns={
                "in_value": "nh4_sheet_in_mg_l",
                "out_value": "nh4_sheet_out_mg_l",
                "removal_rate": "nh4_removal_rate",
            }
        ),
        how="left",
        on=key,
    )

    missing_counts = {
        column: int(water[column].isna().sum())
        for column in ["tn_in_mg_l", "tn_out_mg_l", "tn_removal_rate", "cod_in_mg_l"]
    }
    for column, count in missing_counts.items():
        if count:
            issues.append(
                QualityIssue(
                    "error",
                    "WATER_JOIN_MISSING_VALUE",
                    path.name,
                    "multiple",
                    None,
                    f"按日期和出水序号对齐后，{column} 缺失 {count} 条。",
                )
            )
    counts = {
        "snd_records": int(len(snd)),
        "snd_dates": int(snd["date"].nunique()),
        "tn_outflow_records": int(len(tn)),
        "cod_outflow_records": int(len(cod)),
        "nh4_outflow_records": int(len(nh4)),
    }
    return water, issues, counts


def _metric_map(
    frame: pd.DataFrame,
    start: int,
    label_col: int,
    value_col: int,
    length: int = 8,
) -> dict[str, float]:
    result: dict[str, float] = {}
    for index in range(start, min(start + length, len(frame))):
        label = _clean_text(frame.iat[index, label_col])
        if label:
            result[label] = _number(frame.iat[index, value_col])
    return result


def parse_legacy_our(path: Path) -> tuple[pd.DataFrame, list[QualityIssue]]:
    frame = pd.read_excel(path, sheet_name=0, header=None)
    records: list[dict[str, object]] = []
    issues = _formula_issues(path)
    for start in range(len(frame)):
        if _clean_text(frame.iat[start, 6]) != "OUR(T)":
            continue
        sample_date = canonical_date(frame.iat[start, 9])
        if sample_date is None:
            issues.append(
                QualityIssue(
                    "error",
                    "OUR_BLOCK_WITHOUT_DATE",
                    path.name,
                    "Sheet1",
                    start + 1,
                    "OUR(T) 分块缺少日期。",
                )
            )
            continue
        maxima = _metric_map(frame, start + 1, 0, 2, 7)
        realtime = _metric_map(frame, start, 6, 8, 8)
        records.append(
            {
                "date": sample_date,
                "our_sample_period": "未记录",
                "our_het_max": maxima.get("OUR(het,max)", math.nan),
                "our_aob_max": maxima.get("OUR(A,max)", math.nan),
                "our_nob_max": maxima.get("OUR(N,max)", math.nan),
                "our_het_realtime": realtime.get("OUR(H)", math.nan),
                "our_aob_realtime": realtime.get("OUR(A)", math.nan),
                "our_nob_realtime": realtime.get("OUR(N)", math.nan),
                "our_source_row": start + 1,
            }
        )
    result = pd.DataFrame.from_records(records)
    if not result.empty and result["date"].duplicated().any():
        duplicate_dates = sorted(result.loc[result["date"].duplicated(False), "date"].unique())
        issues.append(
            QualityIssue(
                "warning",
                "DUPLICATE_OUR_DATE",
                path.name,
                "Sheet1",
                None,
                "同日存在多条 OUR 分块：" + "、".join(duplicate_dates),
            )
        )
    return result, issues


def parse_mbr_reactor_our(path: Path) -> tuple[pd.DataFrame, list[QualityIssue]]:
    frame = pd.read_excel(path, sheet_name=0, header=None)
    records: list[dict[str, object]] = []
    issues = _formula_issues(path)
    latest_date: str | None = None
    latest_period = "未记录"

    for index in range(len(frame)):
        cell_date = canonical_date(frame.iat[index, 9])
        if cell_date is not None:
            latest_date = cell_date
            latest_period = "未记录"
        marker = _clean_text(frame.iat[index, 10])
        if marker in {"上午", "下午"}:
            latest_period = marker
        if _clean_text(frame.iat[index, 17]) != "OUR(T)":
            continue
        if latest_date is None:
            issues.append(
                QualityIssue(
                    "error",
                    "OUR_BLOCK_WITHOUT_DATE",
                    path.name,
                    "Sheet1",
                    index + 1,
                    "反应器 OUR(T) 分块之前没有可用日期。",
                )
            )
            continue
        maxima = _metric_map(frame, index + 1, 11, 13, 7)
        realtime = _metric_map(frame, index, 17, 19, 8)
        records.append(
            {
                "date": latest_date,
                "our_sample_period": latest_period,
                "our_het_max": maxima.get("OUR(het,max)", math.nan),
                "our_aob_max": maxima.get("OUR(A,max)", math.nan),
                "our_nob_max": maxima.get("OUR(N,max)", math.nan),
                "our_het_realtime": realtime.get("OUR(H)", math.nan),
                "our_aob_realtime": realtime.get("OUR(A)", math.nan),
                "our_nob_realtime": realtime.get("OUR(N)", math.nan),
                "our_source_row": index + 1,
            }
        )
    result = pd.DataFrame.from_records(records)
    if not result.empty:
        for sample_date, count in result["date"].value_counts().items():
            if count > 1:
                issues.append(
                    QualityIssue(
                        "info",
                        "MULTIPLE_REACTOR_OUR_SAME_DATE",
                        path.name,
                        "Sheet1",
                        None,
                        f"{sample_date} 有 {count} 条反应器 OUR（上午/下午记录应分别保留）。",
                    )
                )
    return result, issues


def _date_alignment_issues(
    water: pd.DataFrame,
    our: pd.DataFrame,
    water_source: str,
    our_source: str,
) -> list[QualityIssue]:
    water_dates = set(water["date"].dropna().astype(str))
    our_dates = set(our["date"].dropna().astype(str))
    issues: list[QualityIssue] = []
    for sample_date in sorted(water_dates - our_dates):
        records = int((water["date"] == sample_date).sum())
        issues.append(
            QualityIssue(
                "warning",
                "WATER_DATE_WITHOUT_OUR",
                water_source,
                "SND",
                None,
                f"{sample_date} 有 {records} 条水质/SND 记录，但没有同日 OUR；未自动改成前一天。",
            )
        )
    for sample_date in sorted(our_dates - water_dates):
        records = int((our["date"] == sample_date).sum())
        issues.append(
            QualityIssue(
                "info",
                "OUR_DATE_WITHOUT_WATER",
                our_source,
                "Sheet1",
                None,
                f"{sample_date} 有 {records} 条 OUR，但没有对应的 SND 记录。",
            )
        )
    return issues


def _build_model_draft(
    water: pd.DataFrame,
    our: pd.DataFrame,
    source_dataset: str,
) -> pd.DataFrame:
    joined = water.merge(our, how="left", on="date")
    draft = pd.DataFrame(
        {
            DATE_COL: joined["date"],
            H_MAX_COL: joined["our_het_max"],
            A_MAX_COL: joined["our_aob_max"],
            N_MAX_COL: joined["our_nob_max"],
            TEMP_COL: np.nan,
            AERATION_COL: np.nan,
            SND_COL: joined["snd_rate"],
            REMOVAL_COL: joined["tn_removal_rate"],
            TN_IN_COL: joined["tn_in_mg_l"],
            COD_IN_COL: joined["cod_in_mg_l"],
            H_LIVE_COL: joined["our_het_realtime"],
            A_LIVE_COL: joined["our_aob_realtime"],
            N_LIVE_COL: joined["our_nob_realtime"],
            TN_OUT_COL: joined["tn_out_mg_l"],
            "source_dataset": source_dataset,
            "source_sample_index": joined["source_sample_index"],
        }
    )
    return draft[MODEL_DRAFT_COLUMNS]


def model_row_audit(frame: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Normalize model inputs and label every source row as included/excluded."""
    missing_columns = [column for column in REQUIRED_MODEL_COLUMNS if column not in frame]
    if missing_columns:
        raise ValueError("缺少必要字段：" + "、".join(missing_columns))

    normalized = frame.copy()
    normalized[DATE_COL] = (
        normalized[DATE_COL].astype("string").str.strip().replace("", pd.NA)
    )
    for column in MODEL_NUMERIC_COLUMNS:
        if column in normalized:
            normalized[column] = pd.to_numeric(normalized[column], errors="coerce")

    audit_rows: list[dict[str, object]] = []
    for position, (_, row) in enumerate(normalized.iterrows(), start=2):
        missing = [column for column in REQUIRED_MODEL_COLUMNS if pd.isna(row[column])]
        invalid: list[str] = []
        if not missing:
            if row[H_MAX_COL] < 0:
                invalid.append(f"{H_MAX_COL}<0")
            if row[A_MAX_COL] < 0:
                invalid.append(f"{A_MAX_COL}<0")
            if row[N_MAX_COL] < 0:
                invalid.append(f"{N_MAX_COL}<0")
            if row[H_LIVE_COL] < 0:
                invalid.append(f"{H_LIVE_COL}<0")
            if row[TN_IN_COL] <= 0:
                invalid.append(f"{TN_IN_COL}<=0")
            if row[COD_IN_COL] < 0:
                invalid.append(f"{COD_IN_COL}<0")
            if row[AERATION_COL] <= 0:
                invalid.append(f"{AERATION_COL}<=0")
            if not 0 <= row[REMOVAL_COL] <= 1:
                invalid.append(f"{REMOVAL_COL}不在0–1")
            if not 0 <= row[SND_COL] <= 1:
                invalid.append(f"{SND_COL}不在0–1")
        reasons = []
        if missing:
            reasons.append("缺失:" + "、".join(missing))
        if invalid:
            reasons.append("越界:" + "、".join(invalid))
        audit_rows.append(
            {
                "CSV行号": position,
                DATE_COL: row[DATE_COL],
                "训练状态": "排除" if reasons else "纳入",
                "排除原因": "；".join(reasons),
            }
        )
    return normalized, pd.DataFrame.from_records(audit_rows)


def validate_model_frame(frame: pd.DataFrame) -> tuple[dict[str, object], list[QualityIssue]]:
    issues: list[QualityIssue] = []
    missing_columns = [column for column in REQUIRED_MODEL_COLUMNS if column not in frame]
    if missing_columns:
        issues.append(
            QualityIssue(
                "error",
                "MISSING_REQUIRED_COLUMNS",
                "model_input",
                "csv",
                None,
                "缺少必要字段：" + "、".join(missing_columns),
            )
        )
        return {
            "train_ready": False,
            "rows": int(len(frame)),
            "date_groups": 0,
            "missing_columns": missing_columns,
        }, issues

    normalized, row_audit = model_row_audit(frame)

    missing_values = {
        column: int(normalized[column].isna().sum()) for column in REQUIRED_MODEL_COLUMNS
    }
    for column, count in missing_values.items():
        if count:
            issues.append(
                QualityIssue(
                    "warning",
                    "MISSING_REQUIRED_VALUES",
                    "model_input",
                    "csv",
                    None,
                    f"{column} 缺失 {count} 条；这些行将在训练前完整案例排除。",
                )
            )

    invalid_masks = {
        H_MAX_COL: normalized[H_MAX_COL] < 0,
        A_MAX_COL: normalized[A_MAX_COL] < 0,
        N_MAX_COL: normalized[N_MAX_COL] < 0,
        H_LIVE_COL: normalized[H_LIVE_COL] < 0,
        TN_IN_COL: normalized[TN_IN_COL] <= 0,
        COD_IN_COL: normalized[COD_IN_COL] < 0,
        AERATION_COL: normalized[AERATION_COL] <= 0,
        REMOVAL_COL: ~normalized[REMOVAL_COL].between(0, 1),
        SND_COL: ~normalized[SND_COL].between(0, 1),
    }
    invalid_counts: dict[str, int] = {}
    for column, mask in invalid_masks.items():
        count = int((mask & normalized[column].notna()).sum())
        invalid_counts[column] = count
        if count:
            issues.append(
                QualityIssue(
                    "warning",
                    "OUT_OF_RANGE_MODEL_VALUE",
                    "model_input",
                    "csv",
                    None,
                    f"{column} 有 {count} 条超出允许范围；这些行将在训练前排除。",
                )
            )

    complete = normalized.dropna(subset=REQUIRED_MODEL_COLUMNS)
    included_mask = row_audit["训练状态"].eq("纳入").to_numpy()
    valid = normalized.loc[included_mask].copy()
    date_groups = int(valid[DATE_COL].nunique())
    if len(valid) < 40:
        issues.append(
            QualityIssue(
                "error",
                "TOO_FEW_VALID_ROWS",
                "model_input",
                "csv",
                None,
                f"完整且范围有效的记录只有 {len(valid)} 条，V4 至少需要 40 条。",
            )
        )
    if date_groups < 5:
        issues.append(
            QualityIssue(
                "error",
                "TOO_FEW_DATE_GROUPS",
                "model_input",
                "csv",
                None,
                f"完整有效日期只有 {date_groups} 个，按日期 5 折至少需要 5 个。",
            )
        )

    summary = {
        "train_ready": not any(issue.severity == "error" for issue in issues),
        "rows": int(len(normalized)),
        "complete_required_rows": int(len(complete)),
        "valid_rows": int(len(valid)),
        "excluded_rows": int((~included_mask).sum()),
        "date_groups": date_groups,
        "missing_columns": [],
        "missing_required_values": missing_values,
        "invalid_value_counts": invalid_counts,
    }
    return summary, issues


def audit_sources(
    *,
    legacy_water_path: Path,
    legacy_our_path: Path,
    current_water_path: Path,
    mbr_our_path: Path,
    output_dir: Path,
) -> dict[str, object]:
    paths = [legacy_water_path, legacy_our_path, current_water_path, mbr_our_path]
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("未找到输入文件：" + "、".join(missing))

    legacy_water, issues_a, legacy_counts = parse_water_quality(
        legacy_water_path, legacy=True
    )
    legacy_our, issues_b = parse_legacy_our(legacy_our_path)
    current_water, issues_c, current_counts = parse_water_quality(
        current_water_path, legacy=False
    )
    current_our, issues_d = parse_mbr_reactor_our(mbr_our_path)

    issues = issues_a + issues_b + issues_c + issues_d
    issues.extend(
        _date_alignment_issues(
            legacy_water, legacy_our, legacy_water_path.name, legacy_our_path.name
        )
    )
    issues.extend(
        _date_alignment_issues(
            current_water, current_our, current_water_path.name, mbr_our_path.name
        )
    )

    legacy_draft = _build_model_draft(legacy_water, legacy_our, "legacy_batch")
    current_draft = _build_model_draft(current_water, current_our, "current_batch")
    model_draft = pd.concat([legacy_draft, current_draft], ignore_index=True)
    validation, validation_issues = validate_model_frame(model_draft)
    issues.extend(validation_issues)

    output_dir.mkdir(parents=True, exist_ok=True)
    draft_path = output_dir / "model_input_draft.csv"
    fill_template_path = output_dir / "model_input_fill_template.csv"
    issues_path = output_dir / "quality_issues.csv"
    report_path = output_dir / "audit_report.json"
    legacy_our_path_out = output_dir / "legacy_our_blocks.csv"
    current_our_path_out = output_dir / "current_reactor_our_blocks.csv"

    available_required = [
        column
        for column in REQUIRED_MODEL_COLUMNS
        if column not in {TEMP_COL, AERATION_COL}
    ]
    fill_template = model_draft.dropna(subset=available_required).copy()
    model_draft.to_csv(draft_path, index=False, encoding="utf-8-sig")
    fill_template.to_csv(fill_template_path, index=False, encoding="utf-8-sig")
    legacy_our.to_csv(legacy_our_path_out, index=False, encoding="utf-8-sig")
    current_our.to_csv(current_our_path_out, index=False, encoding="utf-8-sig")
    issue_frame = pd.DataFrame([issue.as_dict() for issue in issues])
    issue_frame.to_csv(issues_path, index=False, encoding="utf-8-sig")

    severity_counts = defaultdict(int)
    code_counts = defaultdict(int)
    for issue in issues:
        severity_counts[issue.severity] += 1
        code_counts[issue.code] += 1

    report: dict[str, Any] = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": "train_ready" if validation["train_ready"] else "needs_completion",
        "source_files": {path.name: _workbook_metadata(path) for path in paths},
        "record_counts": {
            "legacy_water": legacy_counts,
            "legacy_our_blocks": int(len(legacy_our)),
            "legacy_our_dates": int(legacy_our["date"].nunique()),
            "current_water": current_counts,
            "current_reactor_our_blocks": int(len(current_our)),
            "current_reactor_our_dates": int(current_our["date"].nunique()),
            "model_draft_rows": int(len(model_draft)),
            "water_our_exact_match_rows": int(len(fill_template)),
        },
        "model_input_validation": validation,
        "issue_counts_by_severity": dict(sorted(severity_counts.items())),
        "issue_counts_by_code": dict(sorted(code_counts.items())),
        "outputs": {
            "model_input_draft": str(draft_path.resolve()),
            "model_input_fill_template": str(fill_template_path.resolve()),
            "quality_issues": str(issues_path.resolve()),
            "legacy_our_blocks": str(legacy_our_path_out.resolve()),
            "current_reactor_our_blocks": str(current_our_path_out.resolve()),
            "audit_report": str(report_path.resolve()),
        },
        "notes": [
            "日期仅按同日精确对齐；程序不会把相邻日期的 OUR 自动映射到水质记录。",
            "负硝氮/亚硝氮结果未自动改成零，应结合方法 LOQ 另行编码。",
            "源文件缺失的温度与曝气量保持为空，必须由授权运行记录补齐。",
        ],
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def read_model_csv(path: Path) -> pd.DataFrame:
    errors: list[str] = []
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            frame = pd.read_csv(path, encoding=encoding, dtype={DATE_COL: "string"})
            frame.columns = frame.columns.astype(str).str.strip()
            return frame
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise ValueError("CSV 编码读取失败：" + "；".join(errors))
